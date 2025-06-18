import json
import csv
from openpyxl import load_workbook


class Utilities:

    @staticmethod
    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        rowcount = sh.max_row
        colcount = sh.max_column

        for i in range(2, rowcount + 1):
            row = []
            for j in range(1, colcount + 1):
                row.append(sh.cell(i, j).value)
            datalist.append(row)
        return datalist

    @staticmethod
    def read_data_from_csv(file_name):
        datalist = []
        with open(file_name, "r", newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip header
            for rows in reader:
                datalist.append(rows)
        return datalist

    @staticmethod
    def read_data_from_json(file_name):
        with open(file_name, 'r') as file:
            data = json.load(file)
        return [tuple(item.values()) for item in data]
